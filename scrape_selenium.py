from fake_useragent import UserAgent
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import random
import time
import sys
import os
from openpyxl import Workbook, load_workbook

# 自動根據你的Chrome下載正確的版本
def init_driver():
    driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))
    return driver
 
# 解析電影列表頁面，並抓出電影詳細資訊的 url
def fetch_detail_urls(driver,url):
    try:
        detail_urls = []
        driver.get(url)
        
        try:
            elements = WebDriverWait(driver, 10).until(
                EC.presence_of_all_elements_located((By.CSS_SELECTOR, "a[href^='/detail']"))
            )
        except TimeoutException:
            return detail_urls
        
        for element in elements:
            detail_url = element.get_attribute('href')
            detail_urls.append(detail_url)
            
        detail_urls = list(dict.fromkeys(detail_urls))    
        
        time.sleep(random.uniform(3, 5)) 
        return detail_urls
    except Exception as e:
        error_line = sys.exc_info()[-1].tb_lineno
        print(f"錯誤發生在第 {error_line} 行：\n{str(e)}")
        return None

# 解析電影詳細資訊頁面，並抓出電影詳細資訊頁面中需要的資料
def fetch_detail_data(driver, url):
    try:
        driver.get(url)
        time.sleep(random.uniform(2, 4))
        
        detail_data = []
        
        # 名稱
        name = driver.find_element(By.XPATH, '//*[@id="detail"]/div[1]/div/div/div[1]/div/div[2]/a/h2').text
        detail_data.append(name)
        
        # 評分
        score = driver.find_element(By.CLASS_NAME, 'score').text
        detail_data.append(score)
        
        # 時長
        minute = driver.find_element(By.XPATH, '//*[@id="detail"]/div[1]/div/div/div[1]/div/div[2]/div[2]/span[3]').text
        detail_data.append(minute)
        
        # 簡介
        drama = driver.find_element(By.XPATH, '//*[@id="detail"]/div[1]/div/div/div[1]/div/div[2]/div[4]/p').text
        detail_data.append(drama)
        
        time.sleep(random.uniform(3, 5)) 
          
        return detail_data
    except Exception as e:
        error_line = sys.exc_info()[-1].tb_lineno
        print(f"錯誤發生在第 {error_line} 行：\n{str(e)}")
        return None

# 將資料寫入 Excel 中 
def write_to_excel(data, file_name="scrape_selenium.xlsx"):
    if os.path.exists(file_name):
        workbook = load_workbook(file_name)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["電影名稱", "電影評分", "電影時長", "電影簡介"])
    
    duplicate_found = False   
    for row in sheet.iter_rows(values_only=True):
        if row[0] == data[0]:
            duplicate_found = True
            break
        
    if not duplicate_found:
        sheet.append(data)
        
    workbook.save(file_name)
        
    
if __name__ == "__main__":
    driver = init_driver()
    try:
        for page in range(1,999):
            list_url = f"https://spa1.scrape.center/page/{page}"
            print(f"正在爬取電影數據網站第 {page} 頁")
            detail_urls = fetch_detail_urls(driver, list_url)
            if detail_urls != [] and detail_urls is not None:
                for detail_url in detail_urls:
                    print(f"正在爬取電影數據網站 {detail_url}")
                    detail_data = fetch_detail_data(driver, detail_url)
                    write_to_excel(detail_data)
            else:
                print(f"電影數據網站第 {page-1} 頁是最後一頁囉")
                break
    finally:
        driver.quit()